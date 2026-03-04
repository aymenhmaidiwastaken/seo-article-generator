"""Excel export using openpyxl."""

import logging
import os
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .extractor import ArticleData

logger = logging.getLogger("article_crawler")

# Column definitions
COLUMNS = [
    ("Title", 40),
    ("Article Content", 80),
    ("Source URL", 50),
    ("Date Published", 15),
    ("Author", 20),
    ("Word Count", 12),
    ("Meta Description", 50),
    ("Keywords Found", 30),
    ("Headings", 50),
    ("Source Domain", 25),
    ("Rewritten Title", 40),
    ("Rewritten Content", 80),
    ("Rewrite Status", 15),
    ("Slug", 40),
    ("Category", 15),
]


class ExcelExporter:
    """Exports articles to a formatted Excel file."""

    def __init__(self, output_path: str = "articles_output.xlsx"):
        self.output_path = output_path

    def export(self, articles: List[ArticleData]) -> str:
        """Export articles to Excel file. Returns the file path."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Articles"

        self._setup_header(ws)
        self._write_data(ws, articles)
        self._apply_formatting(ws)

        # Add a summary sheet
        self._add_summary_sheet(wb, articles)

        # Save
        wb.save(self.output_path)
        logger.info(f"Excel file saved: {self.output_path} ({len(articles)} articles)")
        return self.output_path

    def _setup_header(self, ws):
        """Create and style the header row."""
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, (name, width) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"
        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    def _write_data(self, ws, articles: List[ArticleData]):
        """Write article data to rows."""
        content_alignment = Alignment(vertical="top", wrap_text=True)

        for row_idx, article in enumerate(articles, 2):
            values = [
                article.title,
                article.content,
                article.url,
                article.date,
                article.author,
                article.word_count,
                article.meta_description,
                ", ".join(article.keywords_found) if article.keywords_found else "",
                "\n".join(article.headings) if article.headings else "",
                article.source_domain,
                article.rewritten_title,
                article.rewritten_content,
                article.rewrite_status,
                article.slug,
                article.category,
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = content_alignment

            # Color code rewrite status
            status_cell = ws.cell(row=row_idx, column=13)
            if article.rewrite_status == "completed":
                status_cell.fill = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                )
            elif article.rewrite_status == "failed":
                status_cell.fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )
            elif article.rewrite_status == "pending":
                status_cell.fill = PatternFill(
                    start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                )

    def _apply_formatting(self, ws):
        """Apply final formatting touches."""
        # Set row height for data rows
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 60

    def _add_summary_sheet(self, wb: Workbook, articles: List[ArticleData]):
        """Add a summary/stats sheet."""
        ws = wb.create_sheet("Summary")

        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14, color="2F5496")

        ws.cell(row=1, column=1, value="Crawl Summary").font = title_font
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 40

        stats = [
            ("Total Articles", len(articles)),
            ("Average Word Count", int(sum(a.word_count for a in articles) / max(len(articles), 1))),
            ("Min Word Count", min((a.word_count for a in articles), default=0)),
            ("Max Word Count", max((a.word_count for a in articles), default=0)),
            ("", ""),
            ("Rewrite Status", ""),
            ("  Completed", sum(1 for a in articles if a.rewrite_status == "completed")),
            ("  Failed", sum(1 for a in articles if a.rewrite_status == "failed")),
            ("  Pending", sum(1 for a in articles if a.rewrite_status == "pending")),
            ("  Partial", sum(1 for a in articles if a.rewrite_status == "partial")),
            ("  Timeout", sum(1 for a in articles if a.rewrite_status == "timeout")),
            ("", ""),
            ("Top Sources", ""),
        ]

        # Count articles per domain
        domain_counts = {}
        for a in articles:
            domain_counts[a.source_domain] = domain_counts.get(a.source_domain, 0) + 1
        top_domains = sorted(domain_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        for domain, count in top_domains:
            stats.append((f"  {domain}", count))

        # Count categories
        stats.append(("", ""))
        stats.append(("Categories", ""))
        cat_counts = {}
        for a in articles:
            if a.category:
                cat_counts[a.category] = cat_counts.get(a.category, 0) + 1
        for cat, count in sorted(cat_counts.items(), key=lambda x: x[1], reverse=True):
            stats.append((f"  {cat}", count))

        for row_idx, (label, value) in enumerate(stats, 3):
            cell_a = ws.cell(row=row_idx, column=1, value=label)
            cell_b = ws.cell(row=row_idx, column=2, value=value)
            if not label.startswith("  ") and label:
                cell_a.font = header_font

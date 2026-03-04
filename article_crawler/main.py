"""Main entry point and orchestrator for the article crawler."""

import argparse
import logging
import os
import signal
import sys
import time
from typing import List, Set

from tqdm import tqdm

from .checkpoint import CheckpointManager
from .extractor import ArticleData, ArticleExtractor
from .exporter import ExcelExporter
from .rewriter import OllamaRewriter
from .searcher import MultiSearcher
from .utils import random_delay, setup_logging

logger = logging.getLogger("article_crawler")

# Global state for graceful interrupt handling
_interrupt_requested = False
_current_articles = []
_current_args = None


def _handle_interrupt(signum, frame):
    """Handle Ctrl+C gracefully - save progress and export what we have."""
    global _interrupt_requested
    if _interrupt_requested:
        # Second Ctrl+C = force quit
        print("\nForce quit.")
        sys.exit(1)
    _interrupt_requested = True
    print("\n\nInterrupt received! Saving progress and exporting current articles...")
    print("(Press Ctrl+C again to force quit)")


def parse_args():
    parser = argparse.ArgumentParser(
        description="Article Crawler - Crawl the web for articles and rewrite them with AI",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Crawl articles about keto diet
  python -m article_crawler "keto diet" "keto recipes" "ketogenic lifestyle"

  # Crawl with custom output file and target of 500 articles
  python -m article_crawler "home automation" --output smart_home.xlsx --target 500

  # Crawl without rewriting (just collect articles)
  python -m article_crawler "machine learning" --no-rewrite

  # Resume a previous interrupted crawl
  python -m article_crawler "keto diet" --resume

  # Use specific search engines only
  python -m article_crawler "python programming" --engines duckduckgo google

  # Rewrite an existing Excel file (skip crawling)
  python -m article_crawler --rewrite-only existing_articles.xlsx
        """,
    )

    parser.add_argument(
        "keywords",
        nargs="*",
        help="Keywords/phrases to search for (e.g., 'keto diet' 'weight loss')",
    )
    parser.add_argument(
        "-o", "--output",
        default="articles_output.xlsx",
        help="Output Excel file path (default: articles_output.xlsx)",
    )
    parser.add_argument(
        "-t", "--target",
        type=int,
        default=5000,
        help="Target number of articles to collect (default: 5000)",
    )
    parser.add_argument(
        "--min-words",
        type=int,
        default=300,
        help="Minimum word count to keep an article (default: 300)",
    )
    parser.add_argument(
        "--min-relevance",
        type=float,
        default=0.1,
        help="Minimum relevance score 0-1 to keep an article (default: 0.1)",
    )
    parser.add_argument(
        "--skip-dedup",
        action="store_true",
        help="Skip deduplication checks (allow similar articles)",
    )
    parser.add_argument(
        "--no-rewrite",
        action="store_true",
        help="Skip the rewriting phase, just crawl and export raw articles",
    )
    parser.add_argument(
        "--rewrite-only",
        metavar="EXCEL_FILE",
        help="Skip crawling, rewrite articles from an existing Excel file",
    )
    parser.add_argument(
        "--resume",
        action="store_true",
        help="Resume from last checkpoint if available",
    )
    parser.add_argument(
        "--engines",
        nargs="+",
        default=["duckduckgo", "bing", "reddit", "quora"],
        help="Search engines to use (default: duckduckgo, bing, reddit, quora). Add 'google' if you want to try it (usually gets rate limited).",
    )
    parser.add_argument(
        "--model",
        default="llama3",
        help="Ollama model name for rewriting (default: llama3)",
    )
    parser.add_argument(
        "--ollama-url",
        default="http://localhost:11434",
        help="Ollama server URL (default: http://localhost:11434)",
    )
    parser.add_argument(
        "--results-per-engine",
        type=int,
        default=100,
        help="Max results to fetch per engine per query variation (default: 100)",
    )
    parser.add_argument(
        "--checkpoint-every",
        type=int,
        default=25,
        help="Save checkpoint every N articles (default: 25)",
    )
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="Enable verbose/debug logging",
    )

    return parser.parse_args()


def run_crawl(args):
    """Phase 1: Search and extract articles. Returns (articles, all_urls, processed_urls)."""
    checkpoint = CheckpointManager()

    # Initialize or resume
    job_id = CheckpointManager.generate_job_id(args.keywords)
    all_urls: List[str] = []
    processed_urls: Set[str] = set()
    articles: List[ArticleData] = []

    if args.resume and checkpoint.exists(job_id):
        logger.info("Resuming from checkpoint...")
        data = checkpoint.load(job_id)
        if data:
            all_urls = data["all_urls"]
            processed_urls = data["processed_urls"]
            articles = data["articles"]
            phase = data.get("phase", "crawling")
            logger.info(
                f"Resumed: {len(processed_urls)} URLs processed, "
                f"{len(articles)} articles collected, phase={phase}"
            )
            # If we were already past crawling, skip search+extraction
            if phase in ("rewriting", "complete", "crawl_done"):
                logger.info("Search & extraction already done, skipping to next phase")
                return articles, all_urls, processed_urls
    else:
        # Search phase
        logger.info("=" * 60)
        logger.info("PHASE 1: SEARCHING")
        logger.info("=" * 60)
        logger.info(f"Keywords: {', '.join(args.keywords)}")
        logger.info(f"Engines: {', '.join(args.engines)}")

        searcher = MultiSearcher(engines=args.engines)
        all_urls = searcher.search(
            keywords=args.keywords,
            results_per_engine=args.results_per_engine,
        )

        logger.info(f"Found {len(all_urls)} unique URLs to process")

        # Save initial checkpoint with URLs
        checkpoint.save(job_id, args.keywords, all_urls, processed_urls, articles)

    # Extraction phase
    logger.info("=" * 60)
    logger.info("PHASE 2: EXTRACTING ARTICLES")
    logger.info("=" * 60)

    extractor = ArticleExtractor(
        min_word_count=args.min_words,
        target_keywords=args.keywords,
        min_relevance=args.min_relevance,
        skip_dedup=args.skip_dedup,
    )

    remaining_urls = [u for u in all_urls if u not in processed_urls]
    logger.info(f"URLs remaining: {len(remaining_urls)}")

    with tqdm(
        total=min(len(remaining_urls), args.target - len(articles)),
        desc="Extracting articles",
        unit="url",
    ) as pbar:
        for url in remaining_urls:
            if _interrupt_requested:
                logger.info("Interrupt: stopping extraction")
                break

            if len(articles) >= args.target:
                logger.info(f"Target reached: {args.target} articles")
                break

            processed_urls.add(url)
            article = extractor.extract(url)

            if article:
                articles.append(article)
                pbar.set_postfix(articles=len(articles), refresh=True)

            pbar.update(1)

            # Save checkpoint periodically
            if len(processed_urls) % args.checkpoint_every == 0:
                checkpoint.save(
                    job_id, args.keywords, all_urls, processed_urls, articles
                )

            random_delay(0.2, 0.8)

    # Final checkpoint for crawl phase
    checkpoint.save(
        job_id, args.keywords, all_urls, processed_urls, articles, phase="crawl_done"
    )

    logger.info(f"Extraction complete: {len(articles)} articles from {len(processed_urls)} URLs")

    # Print extraction statistics
    extractor.print_stats()

    return articles, all_urls, processed_urls


def run_rewrite(articles: List[ArticleData], args,
                all_urls: List[str] = None,
                processed_urls: Set[str] = None) -> List[ArticleData]:
    """Phase 2: Rewrite articles using Ollama."""
    logger.info("=" * 60)
    logger.info("PHASE 3: REWRITING WITH AI")
    logger.info("=" * 60)

    if all_urls is None:
        all_urls = []
    if processed_urls is None:
        processed_urls = set()

    rewriter = OllamaRewriter(
        model=args.model,
        ollama_url=args.ollama_url,
        target_keyword=args.keywords[0] if args.keywords else "",
    )

    checkpoint = CheckpointManager()
    job_id = CheckpointManager.generate_job_id(args.keywords)

    # Find articles that haven't been rewritten yet
    to_rewrite = [
        (i, a) for i, a in enumerate(articles) if a.rewrite_status == "pending"
    ]

    logger.info(f"Articles to rewrite: {len(to_rewrite)}")

    with tqdm(total=len(to_rewrite), desc="Rewriting articles", unit="article") as pbar:
        for idx, (orig_idx, article) in enumerate(to_rewrite):
            if _interrupt_requested:
                logger.info("Interrupt: stopping rewrite")
                break

            articles[orig_idx] = rewriter.rewrite(article)
            pbar.update(1)

            status = articles[orig_idx].rewrite_status
            pbar.set_postfix(status=status, refresh=True)

            # Save checkpoint periodically
            if (idx + 1) % 5 == 0:
                checkpoint.save(
                    job_id,
                    args.keywords,
                    all_urls,
                    processed_urls,
                    articles,
                    phase="rewriting",
                    rewrite_index=idx + 1,
                )

            # Small delay between rewrites to avoid overwhelming Ollama
            random_delay(0.5, 1.0)

    # Final checkpoint
    checkpoint.save(
        job_id, args.keywords, all_urls, processed_urls, articles,
        phase="complete", rewrite_index=len(to_rewrite),
    )

    completed = sum(1 for a in articles if a.rewrite_status == "completed")
    logger.info(f"Rewriting complete: {completed}/{len(articles)} successful")
    return articles


def load_articles_from_excel(filepath: str) -> List[ArticleData]:
    """Load articles from an existing Excel file for rewriting."""
    from openpyxl import load_workbook

    if not os.path.exists(filepath):
        logger.error(f"File not found: {filepath}")
        sys.exit(1)

    wb = load_workbook(filepath)
    ws = wb.active
    articles = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        article = ArticleData(
            url=str(row[2] or ""),
            title=str(row[0] or ""),
            content=str(row[1] or ""),
            date=str(row[3] or ""),
            author=str(row[4] or ""),
            word_count=int(row[5] or 0),
            meta_description=str(row[6] or ""),
            keywords_found=str(row[7] or "").split(", ") if row[7] else [],
            headings=str(row[8] or "").split("\n") if row[8] else [],
            source_domain=str(row[9] or ""),
            rewritten_title=str(row[10] or ""),
            rewritten_content=str(row[11] or ""),
            rewrite_status=str(row[12] or "pending"),
            slug=str(row[13] or ""),
            category=str(row[14] or ""),
        )
        articles.append(article)

    logger.info(f"Loaded {len(articles)} articles from {filepath}")
    return articles


def main():
    global _current_articles, _current_args

    args = parse_args()
    _current_args = args
    setup_logging(verbose=args.verbose)

    # Install graceful interrupt handler
    signal.signal(signal.SIGINT, _handle_interrupt)

    start_time = time.time()

    print()
    print("=" * 60)
    print("  ARTICLE CRAWLER v1.0")
    print("  Crawl. Extract. Rewrite. Export.")
    print("=" * 60)
    print()

    articles = []

    # Mode: Rewrite only
    if args.rewrite_only:
        if not args.keywords:
            args.keywords = ["article"]  # fallback for job_id
        articles = load_articles_from_excel(args.rewrite_only)
        articles = run_rewrite(articles, args)
        exporter = ExcelExporter(output_path=args.output)
        exporter.export(articles)

        elapsed = time.time() - start_time
        print(f"\nDone! {len(articles)} articles rewritten in {elapsed:.0f}s")
        print(f"Output: {os.path.abspath(args.output)}")
        return

    # Validate keywords
    if not args.keywords:
        print("Error: Please provide at least one keyword.")
        print("Usage: python -m article_crawler 'keto diet' 'weight loss'")
        sys.exit(1)

    # Phase 1: Crawl
    articles, all_urls, processed_urls = run_crawl(args)
    _current_articles = articles

    if not articles:
        print("\nNo articles found. Try different keywords or engines.")
        sys.exit(1)

    # Phase 2: Rewrite (unless --no-rewrite or interrupted)
    if not args.no_rewrite and not _interrupt_requested:
        articles = run_rewrite(articles, args, all_urls, processed_urls)
        _current_articles = articles

    # Phase 3: Export
    logger.info("=" * 60)
    logger.info("PHASE 4: EXPORTING TO EXCEL")
    logger.info("=" * 60)

    exporter = ExcelExporter(output_path=args.output)
    output_path = exporter.export(articles)

    elapsed = time.time() - start_time

    print()
    print("=" * 60)
    print("  CRAWL COMPLETE")
    print("=" * 60)
    print(f"  Articles collected:  {len(articles)}")
    if not args.no_rewrite:
        completed = sum(1 for a in articles if a.rewrite_status == "completed")
        print(f"  Articles rewritten:  {completed}")
    print(f"  Output file:         {os.path.abspath(output_path)}")
    print(f"  Total time:          {elapsed:.0f}s")
    print("=" * 60)
    print()
